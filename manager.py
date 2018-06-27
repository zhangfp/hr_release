#-*-coding=utf-8-*-
from flask import Flask, request, session, g, redirect, url_for, abort, \
     render_template, flash

from wtforms import StringField,SubmitField,IntegerField,TextField,BooleanField,PasswordField
from wtforms.validators import DataRequired,Required,Length,Email

from flask.ext.wtf import Form
from flask.ext.login import LoginManager,login_required, login_user,logout_user,UserMixin 

from flask.ext.sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash,check_password_hash

import os,subprocess
import time

from werkzeug import secure_filename
from datetime import datetime

# import json
# from pymongo import MongoClient
# from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import numbers
# import collections
import xlrd
import re

# file_name = sys.argv[1]
# print file_name
# 连接数据库
# client=MongoClient('localhost',27017)
# mongodb=client.hr
# hr=mongodb.hr
# hr_keys=mongodb.hr_keys
# project_map_table=mongodb.project_map

##########
currFolder = os.path.dirname(os.path.realpath(__file__));
UPLOAD_FOLDER = os.path.join(currFolder, 'static/uploads')
DOWNLOAD_FOLDER = os.path.join(currFolder, 'static/downloads')
ALLOWED_EXTENSIONS = set(['xlsx', 'xls'])


destFolder = '/data/release/sgonline/'
# destFolder = '/tmp/'
debug_flag = True
host_group_lianyun = 'webservers'
host_group_4399 = '4399_web'
##########
# currFolder = os.path.dirname(os.path.realpath(__file__));
app = Flask(__name__)
app.config.update(dict(
    SECRET_KEY='development key',
    SQLALCHEMY_DATABASE_URI ='sqlite:///' + os.path.join(currFolder, 'data.sqlite'))
    )
# bootstrap = Bootstrap(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
db = SQLAlchemy()
db.init_app(app)

download_filename = ''


def get_excel_value(cell, ctype):
    if ctype == 2 and ctype % 1 == 0:
        cell = int(cell)
    elif ctype == 3:
        date = datetime(*xlrd.xldate_as_tuple(cell, 0))
        cell = date.strftime('%Y-%m-%d')
    elif ctype == 4:
        cell = True if cell == 1 else False
    return cell


def get_col(sheet, rows, cols, col_str):
    for i in range(0, rows):
        for j in range(0, cols):
            cell_value = sheet.cell_value(i, j)
            if cell_value == col_str:
                return j


def import_one_team(file_name, start_time, end_time):
    data=xlrd.open_workbook(file_name)
    sheet0 = data.sheets()[0]
    sheet1 = data.sheets()[1]
    # project_map = project_map_table.find_one()
    sheet0_nrows = sheet0.nrows
    sheet0_ncols = sheet0.ncols
    sheet1_nrows = sheet1.nrows
    sheet1_ncols = sheet1.ncols
    timed_delivery_num = 0
    planed_delivery_num = 0
    test_delivery_num = 0
    temp_delivery_num = 0
    test_pass_num = 0
    test_num = 0
    formal_delivery_rate = 0
    test_pass_rate = 0

    planed_prepublish_time_col = get_col(sheet0, sheet0_nrows, sheet0_ncols, u"计划预发布时间")
    real_prepublish_time_col = get_col(sheet0, sheet0_nrows, sheet0_ncols, u"实际预发布时间")
    planed_publish_time_col = get_col(sheet0, sheet0_nrows, sheet0_ncols, u"计划交付现场时间")
    real_publish_time_col = get_col(sheet0, sheet0_nrows, sheet0_ncols, u"实际交付现场时间")
    required_col = get_col(sheet0, sheet0_nrows, sheet0_ncols, u"要求")
    test_result_col = get_col(sheet1, sheet1_nrows, sheet1_ncols, u"测试结果")
    prepublish_time_col = get_col(sheet1, sheet1_nrows, sheet1_ncols, u"预发布时间")
    publish_time_col = get_col(sheet1, sheet1_nrows, sheet1_ncols, u"正式发布时间")

    for i in range(1, sheet0_nrows):
        # print table.row_values(i)
        ctype = sheet0.cell(i,planed_prepublish_time_col).ctype
        cell = sheet0.cell_value(i, planed_prepublish_time_col)
        planed_prepublish_time = get_excel_value(cell, ctype)

        ctype = sheet0.cell(i,real_prepublish_time_col).ctype
        cell = sheet0.cell_value(i, real_prepublish_time_col)
        real_prepublish_time = get_excel_value(cell, ctype)

        ctype = sheet0.cell(i,planed_publish_time_col).ctype
        cell = sheet0.cell_value(i, planed_publish_time_col)
        planed_publish_time = get_excel_value(cell, ctype)

        ctype = sheet0.cell(i,real_publish_time_col).ctype
        cell = sheet0.cell_value(i, real_publish_time_col)
        real_publish_time = get_excel_value(cell, ctype)

        ctype = sheet0.cell(i,required_col).ctype
        cell = sheet0.cell_value(i, required_col)
        required = get_excel_value(cell, ctype)

        # 准时交付版本总数
        if required == u"正式发布":
            mat = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", real_publish_time)
            if mat and mat.group(0) == real_publish_time:
                mat2 = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", planed_publish_time)
                if mat2 and mat2.group(0) == planed_publish_time:
                    if start_time <= planed_publish_time <= end_time \
                            and real_publish_time <= planed_publish_time:
                        timed_delivery_num += 1
                else:
                    if start_time <= real_publish_time <= end_time:
                        timed_delivery_num += 1
                        # print "planed_publish_time:", planed_publish_time
                        # print "real_publish_time:", real_publish_time
        # else:
        #     if start_time <= planed_publish_time <= end_time \
        #             and real_publish_time <= planed_publish_time:
        #         timed_delivery_num += 1

        # 计划交付版本总数
        if start_time <= planed_publish_time <= end_time and required == u"正式发布":
            mat2 = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", planed_publish_time)
            if mat2 and mat2.group(0) == planed_publish_time:
                planed_delivery_num += 1
        elif start_time <= planed_prepublish_time <= end_time and required == u"预发布":
            mat2 = re.match(r"(\d{4}-\d{1,2}-\d{1,2})", planed_prepublish_time)
            if mat2 and mat2.group(0) == planed_prepublish_time:
                planed_delivery_num += 1

        # 测试发布版本总数
        if start_time <= real_publish_time <= end_time \
                and required == u"正式发布" \
                and re.match(r"(\d{4}-\d{1,2}-\d{1,2})", real_publish_time):
            test_delivery_num += 1

        # 临时版本数
        if start_time <= real_prepublish_time <= end_time \
                and required == u"预发布" \
                and re.match(r"(\d{4}-\d{1,2}-\d{1,2})", real_prepublish_time):
            temp_delivery_num += 1
    
    if test_result_col < sheet1_ncols and prepublish_time_col <= sheet1_ncols and publish_time_col < sheet1_ncols:
        for i in range(1, sheet1_nrows):
            # 测试版本总数
            ctype = sheet1.cell(i, test_result_col).ctype
            cell = sheet1.cell_value(i, test_result_col)
            test_result = get_excel_value(cell, ctype)

            ctype = sheet1.cell(i, prepublish_time_col).ctype
            cell = sheet1.cell_value(i, prepublish_time_col)
            prepublish_time = get_excel_value(cell, ctype)

            ctype = sheet1.cell(i, publish_time_col).ctype
            cell = sheet1.cell_value(i, publish_time_col)
            publish_time = get_excel_value(cell, ctype)
#            print "test_result:",test_result
            if test_result == 'ok' or test_result == 'OK' or test_result == 'nok' or test_result == 'NOK':
                if start_time <= prepublish_time <= end_time:
                    test_num += 1
            if test_result == 'ok' or test_result == 'OK':
                if start_time <= prepublish_time <= end_time:
                    test_pass_num += 1
                # elif (test_result == 'ok' or test_result == 'OK') \
                #         and start_time <= publish_time <= end_time:
                #     test_num += 1
    try:
        # 正式版本准时交付率
        formal_delivery_rate = timed_delivery_num * 1.0 / test_delivery_num
        # 版本通过率
        test_pass_rate = test_pass_num * 1.0 / test_num
    except ZeroDivisionError, e:
        print e.message
    finally:
        pass
    
    print "timed_delivery_num:", timed_delivery_num
    print "planed_delivery_num:", planed_delivery_num
    print "test_delivery_num:", test_delivery_num
    print "temp_delivery_num:", temp_delivery_num
    print "test_num:", test_num
    print "formal_delivery_rate:", formal_delivery_rate
    print "test_pass_rate:", test_pass_rate
    filebasename = os.path.basename(file_name)
    ret = [filebasename, timed_delivery_num, planed_delivery_num,
           test_delivery_num, temp_delivery_num,
           test_num, formal_delivery_rate, test_pass_rate]
    return ret


class User(UserMixin,db.Model):
    print "into class User..............."
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(64), unique=True, index=True)
    username = db.Column(db.String(64), unique=True, index=True)
    # 密码为128的hash值
    password_hash = db.Column(db.String(128))
    # 增加一列判断是否为管理员
    is_admin = db.Column(db.Boolean,default = False)
    role_id = db.Column(db.Integer, db.ForeignKey('roles.id'))

    def __repr__(self):
        return '<User %r>' % self.username

    # password设置property使密码不可直接读，verify_password()判断密码是否正确。
    @property
    def password(self):
        raise AttributeError('password is not a readable attribute')

    @password.setter
    def password(self,password):
        self.password_hash = generate_password_hash(password)

    def verify_password(self,password):
        return check_password_hash(self.password_hash, password)

    def __init__(self, **kwargs):
        super(User, self).__init__(**kwargs)
        if self.role is None:
            if self.is_admin:
                self.role = Role.query.filter_by(name="Administrator").first()
            if self.role is None:
                self.role = Role.query.filter_by(name="User").first()

    # 用户权限验证
    def can(self, permissions):
        return self.role is not None and \
            (self.role.permissions & permissions) == permissions

    def is_administrator(self):
        print "is_admin starting...."
        return self.can(Permission.ADMINISTER)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class BaseForm(Form):
    LANGUAGES = ['zh']


class LianyunForm(BaseForm):
    main_program = BooleanField(u'发主程序', validators = [DataRequired()])
    main_xml = BooleanField(u'发主配置', validators = [DataRequired()])
    submit = SubmitField(u'提交')


class LoginForm(BaseForm):
    email = StringField(u'邮箱', validators=[Required(), Length(1, 64),Email()])
    password = PasswordField(u'密码', validators=[Required()])
    # remember_me = BooleanField(u'下次自动登录')
    submit = SubmitField(u'登陆')


def subprores(command,success_res = u'Execution OK'):
    """执行command，返回状态码
    """
    try:
        # retcode = subprocess.call(command,shell=True)
        ret = subprocess.Popen(command,stdout=subprocess.PIPE,shell=True)
        retMessage = ret.stdout.read()
        retcode = ret.poll()
        if retcode == 0 or retcode == None:
            print retMessage
            return success_res + u' 成功'
        else:
            print "Child returned",retcode
            print retMessage
            return success_res + u' 失败'
    except OSError as e:
        print >> sys.stderr, "Execution failed", e
        return False

#def check_message(res):
#    if u'失败' in res:
#        flash(u'请联系运维处理')
#        return True


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        print "request.files:", request.files
        files = request.files.getlist("file")
        filenames = []
        for file in files:
            print "file:", file
            print "file.filename", file.filename
            filenames.append(file.filename)
            if file and allowed_file(file.filename):
                # filename = secure_filename(file.filename)
                filename = file.filename
                now_time = datetime.now().strftime('%Y-%m-%d')
                str_now_time = str(now_time)
                upload_file_path = UPLOAD_FOLDER
                if not os.path.exists(upload_file_path):
                    os.makedirs(upload_file_path)
                # int(time.time())
                file.save(os.path.join(upload_file_path, filename))
                # import_file_name  = upload_file_path + '/' + filename
                # import_one(import_file_name)
                # return redirect(url_for('uploaded_file', filename=filename))
        print "filenames:",filenames
        flash(u"上传完成")
        return redirect(url_for('index'))

    # return render_template('index_upload.html')
    # return render_template('team.html')
    return redirect(url_for('team'))


@app.route('/team', methods=['GET', 'POST'])
def team():
    if request.method == 'POST':
        print "request.files:", request.files
        files = request.files.getlist("file")
        # request.
        filenames = []
        upload_file_path = UPLOAD_FOLDER + "/team"
        if os.path.exists(upload_file_path):
            for team_file in os.listdir(upload_file_path):
                os.remove(os.path.join(upload_file_path, team_file))
        else:
            os.makedirs(upload_file_path)

        for file in files:
            print "file:", file
            print "file.filename", file.filename
            filenames.append(file.filename)
            if file and allowed_file(file.filename):
                # filename = secure_filename(file.filename)
                filename = file.filename
                now_time = datetime.now().strftime('%Y-%m-%d')
                str_now_time = str(now_time)
                # int(time.time())
                file.save(os.path.join(upload_file_path, filename))
                # import_file_name  = upload_file_path + '/' + filename
                # import_one(import_file_name)
                # return redirect(url_for('uploaded_file', filename=filename))
        print "filenames:",filenames
        flash(u"上传完成")
        return redirect(url_for('team'))
    global download_filename
    download_filename = u"downloads/" + download_filename
    return render_template('team.html', download_file_name=download_filename)


@app.route('/start_team_statistics', methods=['GET', 'POST'])
def start_team_statistics():
    start_time = request.values['starttime']
    end_time = request.values['endtime']
    print "start time: ", start_time
    print "end time:", end_time
    if not start_time or not end_time:
        flash(u"日期为空，请填写正确日期。")
        return redirect(url_for('team'))

    team_dir = UPLOAD_FOLDER + "/team"
    team_list = []
    for team_file in os.listdir(team_dir):
        print team_file
        team_list.append(import_one_team(UPLOAD_FOLDER + "/team/" + team_file, start_time, end_time))
        print team_list
    wb = Workbook()
    ws = wb.active
    ws.title = "team"
    title = [u"团队", u"准时交付版本总数（正式版本only）", u"计划交付版本总数", u"测试发布版本总数（正式版本正式发布数)",
             u"临时版本数", u"测试版本总数（正式版本only）",
             u"正式版本准时交付率", u"版本通过率"]
    ws.append(title)
    for row in team_list:
        ws.append(row)
        ws.cell(row=ws.max_row, column=ws.max_column-1).number_format = numbers.FORMAT_PERCENTAGE_00
        ws.cell(row=ws.max_row, column=ws.max_column).number_format = numbers.FORMAT_PERCENTAGE_00
    if not os.path.exists(DOWNLOAD_FOLDER):
        os.mkdir(DOWNLOAD_FOLDER)
    global download_filename
    now_time = datetime.now().strftime('%Y-%m-%d %H%M%S')
    str_now_time = str(now_time)
    download_filename = u"项目及团队情况统计原始数据表" + start_time + "_" + end_time + "_" + str_now_time + ".xlsx"
    wb.save(DOWNLOAD_FOLDER + "/" + download_filename)
    flash(u"统计完成，请下载。")

    return redirect(url_for('team'))


@app.route('/project', methods=['GET', 'POST'])
def project():
    if request.method == 'POST':
        print "request.files:", request.files
        files = request.files.getlist("file")
        filenames = []
        for file in files:
            print "file:", file
            print "file.filename", file.filename
            filenames.append(file.filename)
            if file and allowed_file(file.filename):
                # filename = secure_filename(file.filename)
                filename = file.filename
                now_time = datetime.now().strftime('%Y-%m-%d')
                str_now_time = str(now_time)
                upload_file_path = UPLOAD_FOLDER + "/project"
                if not os.path.exists(upload_file_path):
                    os.makedirs(upload_file_path)
                # int(time.time())
                file.save(os.path.join(upload_file_path, filename))
                # import_file_name  = upload_file_path + '/' + filename
                # import_one(import_file_name)
                # return redirect(url_for('uploaded_file', filename=filename))
        print "filenames:",filenames
        flash(u"上传完成")
        return redirect(url_for('project'))

    return render_template('project.html')


@app.route('/start_project_statistics', methods=['GET', 'POST'])
def start_project_statistics():
    # if request.method == 'POST':
    #     print "request.files:", request.files
    #     files = request.files.getlist("file")
    #     filenames = []
    #     for file in files:
    #         print "file:", file
    #         print "file.filename", file.filename
    #         filenames.append(file.filename)
    #         if file and allowed_file(file.filename):
    #             # filename = secure_filename(file.filename)
    #             filename = file.filename
    #             now_time = datetime.now().strftime('%Y-%m-%d')
    #             str_now_time = str(now_time)
    #             upload_file_path = UPLOAD_FOLDER + "/project"
    #             if not os.path.exists(upload_file_path):
    #                 os.makedirs(upload_file_path)
    #             # int(time.time())
    #             file.save(os.path.join(upload_file_path, filename))
    #             # import_file_name  = upload_file_path + '/' + filename
    #             # import_one(import_file_name)
    #             # return redirect(url_for('uploaded_file', filename=filename))
    #     print "filenames:",filenames
    #     flash(u"上传完成")
    #     return redirect(url_for('project'))

    return render_template('project.html')



@app.route('/login',  methods=['GET', "POST"])
def login():
    Form = LoginForm()
    if Form.email.data and Form.password.data:
        user = User.query.filter_by(email=Form.email.data).first()
        print user
        if user is not None and user.verify_password(Form.password.data):
            login_user(user, False)
            return redirect(request.args.get('next') or url_for('index'))
        flash(u'用户名或密码错误，请重新输入。')
    return render_template('login.html', form=Form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


if __name__ == '__main__':
    # app.run(port=80,debug=False,host='0.0.0.0')
    app.run(port=80, debug=True,host='0.0.0.0')

